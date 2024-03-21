import matplotlib.pyplot as plt
from openpyxl import load_workbook

# Load the workbook and select the first sheet
workbook = load_workbook(filename='Performance.xlsx', data_only=True)
first_sheet = workbook.worksheets[0]

# Data extraction from Excel sheet
data = {
    "No Copy": {"Max Depth": [], "RAM Average": [], "RAM Max": [], "CPU Average": [], "CPU Max": []},
    "Copy": {"Max Depth": [], "RAM Average": [], "RAM Max": [], "CPU Average": [], "CPU Max": []},
    "Deep Copy": {"Max Depth": [], "RAM Average": [], "RAM Max": [], "CPU Average": [], "CPU Max": []}
}

# Start reading data from row 4 onwards
for row in first_sheet.iter_rows(min_row=4, values_only=True):
    max_depth_str = row[0]
    if max_depth_str is not None and "max_depth=" in max_depth_str:
        try:
            max_depth = int(max_depth_str.split("=")[1])
            if 4 <= max_depth <= 7:
                for key, (cpu_avg_col, cpu_max_col, ram_avg_col, ram_max_col) in zip(
                    data.keys(), [(1, 2, 3, 4), (5, 6, 7, 8), (9, 10, 11, 12)]
                ):
                    data[key]["Max Depth"].append(max_depth)
                    data[key]["CPU Average"].append(row[cpu_avg_col])
                    data[key]["CPU Max"].append(row[cpu_max_col])
                    data[key]["RAM Average"].append(row[ram_avg_col])
                    data[key]["RAM Max"].append(row[ram_max_col])
        except IndexError:
            # Handle cases where the row does not have all the expected columns
            print(f"Row with max_depth={max_depth} is incomplete and will be skipped.")
            continue

# Plotting function
def plot_combined_data(copy_type, data):
    fig, axs = plt.subplots(1, 2, figsize=(20, 6))  # Create 2 subplots side by side

    # Plot RAM Consumption
    axs[0].plot(data[copy_type]["Max Depth"], data[copy_type]["RAM Average"], label='RAM Average', marker='o')
    axs[0].plot(data[copy_type]["Max Depth"], data[copy_type]["RAM Max"], label='RAM Max', marker='s')
    axs[0].set_title(f'{copy_type} - RAM Consumption vs Max Depth')
    axs[0].set_xlabel('Max Depth')
    axs[0].set_ylabel('RAM Consumption (units)')
    axs[0].legend()
    axs[0].grid(True)

    # Plot CPU Consumption
    axs[1].plot(data[copy_type]["Max Depth"], data[copy_type]["CPU Average"], label='CPU Average', marker='o')
    axs[1].plot(data[copy_type]["Max Depth"], data[copy_type]["CPU Max"], label='CPU Max', marker='s')
    axs[1].set_title(f'{copy_type} - CPU Consumption vs Max Depth')
    axs[1].set_xlabel('Max Depth')
    axs[1].set_ylabel('CPU Consumption (units)')
    axs[1].legend()
    axs[1].grid(True)

    plt.show()

# Generate combined plots for each copy type
for copy_type in data.keys():
    plot_combined_data(copy_type, data)
